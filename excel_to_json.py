from openpyxl import load_workbook
import json
from datetime import datetime, date

FILE = "/Users/connorsweeney/Desktop/East Tampa CRM - v20 - AI Enriched.xlsx"
SHEET = "Leased"

def clean_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return ""
    return value

wb = load_workbook(FILE, data_only=True)
ws = wb[SHEET]

headers = [cell.value for cell in ws[1]]

data = []

for row in ws.iter_rows(min_row=2, values_only=True):
    entry = {}
    for i, value in enumerate(row):
        entry[str(headers[i])] = clean_value(value)
    data.append(entry)

with open("data.json", "w") as f:
    json.dump(data, f, indent=2)

print(f"Exported {len(data)} rows to data.json")
